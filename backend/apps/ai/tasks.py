import os
import logging
import io
from celery import shared_task
from django.conf import settings
from django.core.files.base import ContentFile

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def generate_report_task(self, report_id):
    from apps.reports.models import Report
    from apps.reports.repositories import ReportRepository

    try:
        report = Report.objects.select_related('project', 'created_by').get(id=report_id)
    except Report.DoesNotExist:
        logger.error('Report %s not found', report_id)
        return

    ReportRepository.mark_generating(report)

    try:
        if report.format == 'csv':
            content, filename = _generate_csv(report)
        elif report.format == 'excel':
            content, filename = _generate_excel(report)
        elif report.format == 'pdf':
            content, filename = _generate_pdf(report)
        else:
            raise ValueError(f'Unknown format: {report.format}')

        report.file.save(filename, ContentFile(content), save=False)
        ReportRepository.mark_ready(report, report.file.name, len(content))

        _notify_report_ready(report)
        logger.info('Report %s generated successfully', report_id)

    except Exception as exc:
        logger.error('Report generation failed for %s: %s', report_id, exc)
        ReportRepository.mark_failed(report, str(exc))
        raise self.retry(exc=exc, countdown=60)


def _generate_csv(report):
    import csv
    output = io.StringIO()
    writer = csv.writer(output)

    from apps.seo.repositories import SEORepository
    from datetime import date, timedelta

    date_from = report.date_from or (date.today() - timedelta(days=28))
    date_to = report.date_to or date.today()

    writer.writerow(['Keyword', 'Clicks', 'Impressions', 'CTR', 'Position'])
    keywords = SEORepository.get_top_keywords(report.project, date_from, date_to, limit=1000)
    for row in keywords:
        writer.writerow([
            row['keyword__keyword'],
            row['total_clicks'],
            row['total_impressions'],
            f"{float(row['avg_ctr']):.4f}",
            f"{float(row['avg_position']):.2f}",
        ])

    content = output.getvalue().encode('utf-8-sig')
    filename = f'report_{report.id}_{report.name}.csv'
    return content, filename


def _generate_excel(report):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from apps.seo.repositories import SEORepository
    from datetime import date, timedelta

    date_from = report.date_from or (date.today() - timedelta(days=28))
    date_to = report.date_to or date.today()

    wb = openpyxl.Workbook()

    ws_keywords = wb.active
    ws_keywords.title = 'Keywords'
    headers = ['Keyword', 'Clicks', 'Impressions', 'CTR', 'Position']
    for col, header in enumerate(headers, 1):
        cell = ws_keywords.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    keywords = SEORepository.get_top_keywords(report.project, date_from, date_to, limit=1000)
    for row_idx, row in enumerate(keywords, 2):
        ws_keywords.cell(row=row_idx, column=1, value=row['keyword__keyword'])
        ws_keywords.cell(row=row_idx, column=2, value=row['total_clicks'])
        ws_keywords.cell(row=row_idx, column=3, value=row['total_impressions'])
        ws_keywords.cell(row=row_idx, column=4, value=float(row['avg_ctr']))
        ws_keywords.cell(row=row_idx, column=5, value=float(row['avg_position']))

    ws_pages = wb.create_sheet(title='Pages')
    page_headers = ['URL', 'Clicks', 'Impressions', 'CTR', 'Position']
    for col, header in enumerate(page_headers, 1):
        cell = ws_pages.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    pages = SEORepository.get_top_pages(report.project, date_from, date_to, limit=1000)
    for row_idx, row in enumerate(pages, 2):
        ws_pages.cell(row=row_idx, column=1, value=row['page__url'])
        ws_pages.cell(row=row_idx, column=2, value=row['total_clicks'])
        ws_pages.cell(row=row_idx, column=3, value=row['total_impressions'])
        ws_pages.cell(row=row_idx, column=4, value=float(row['avg_ctr']))
        ws_pages.cell(row=row_idx, column=5, value=float(row['avg_position']))

    output = io.BytesIO()
    wb.save(output)
    content = output.getvalue()
    filename = f'report_{report.id}_{report.name}.xlsx'
    return content, filename


def _generate_pdf(report):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from apps.seo.repositories import SEORepository
    from datetime import date, timedelta

    date_from = report.date_from or (date.today() - timedelta(days=28))
    date_to = report.date_to or date.today()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f'SEO Report: {report.project.name}', styles['Title']))
    elements.append(Paragraph(f'Period: {date_from} to {date_to}', styles['Normal']))
    elements.append(Spacer(1, 20))

    elements.append(Paragraph('Top Keywords', styles['Heading2']))
    keyword_data = [['Keyword', 'Clicks', 'Impressions', 'CTR', 'Position']]
    keywords = SEORepository.get_top_keywords(report.project, date_from, date_to, limit=50)
    for row in keywords:
        keyword_data.append([
            str(row['keyword__keyword'])[:60],
            str(row['total_clicks']),
            str(row['total_impressions']),
            f"{float(row['avg_ctr']):.4f}",
            f"{float(row['avg_position']):.2f}",
        ])

    table = Table(keyword_data, colWidths=[200, 60, 80, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
    ]))
    elements.append(table)

    doc.build(elements)
    content = output.getvalue()
    filename = f'report_{report.id}_{report.name}.pdf'
    return content, filename


def _notify_report_ready(report):
    try:
        from apps.notifications.services import NotificationService
        service = NotificationService()
        service.create_notification(
            recipient=report.created_by,
            sender=None,
            notification_type='report_ready',
            title=f'Report ready: {report.name}',
            body=f'Your {report.format.upper()} report has been generated.',
            action_url=f'/reports/{report.id}',
        )
    except Exception as e:
        logger.error('Failed to send report notification: %s', e)


@shared_task
def run_scheduled_reports_task():
    from apps.reports.repositories import ScheduledReportRepository, ReportRepository

    due_reports = ScheduledReportRepository.get_due_reports()
    for scheduled in due_reports:
        try:
            report = ReportRepository.create(
                project=scheduled.project,
                created_by=scheduled.created_by,
                name=f'{scheduled.name} (Auto)',
                format=scheduled.format,
                config=scheduled.config,
            )
            generate_report_task.delay(str(report.id))
            ScheduledReportRepository.update_next_run(scheduled)
            logger.info('Queued scheduled report: %s', scheduled.name)
        except Exception as e:
            logger.error('Failed to queue scheduled report %s: %s', scheduled.name, e)